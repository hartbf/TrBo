# =============================================================================
# BOT CORE FUNCTIONALITY
# =============================================================================
import time
import csv
import openpyxl
import asyncio
import importlib
import nest_asyncio
from openpyxl.styles import PatternFill
from threading import Thread, Lock
from datetime import datetime
from binance.client import Client
from binance import ThreadedWebsocketManager
from binance.enums import *
from config import *
from api_utils import api_call, get_current_price, bot_status_check
from api_utils import fmt_qty_safe
from general_utils import check_all_packages
from logging_setup import log_info
import subprocess
import os


def sync_windows_time():
    if os.name != 'nt':
        print("Zeit-Sync nur unter Windows verfügbar.")
        return
    try:
        subprocess.run('net stop w32time', shell=True, check=True)
        subprocess.run('w32tm /unregister', shell=True, check=True)
        subprocess.run('w32tm /register', shell=True, check=True)
        subprocess.run('net start w32time', shell=True, check=True)
        subprocess.run('w32tm /resync', shell=True, check=True)
        print("Systemzeit synchronisiert.")
    except Exception as e:
        print("Fehler bei Zeit-Synchronisation:", e)
class GridTradingBot:
    def __init__(self):
        self.client = Client(API_KEY, API_SECRET)
        self.running = False
        self.bot_active = False
        self.positions_btc = 0.0
        self.usdc_balance = 0.0
        self.btc_balance = 0.0
        self.net_profit_usdc = 0.0
        self.total_invested_usdc = 0.0
        self.last_price = None
        self.order_history = []
        self.lock = Lock()
        self.twm = None
        self.last_stream_time = time.time()
        self.reconnect_delay = 5
        self.open_orders = {}
        self.initial_btc_balance = 0.01
        self.initial_usdc_balance = 1000.0
        self.btc_balance = self.initial_btc_balance
        self.usdc_balance = self.initial_usdc_balance
        self.fee_rate = 0.0001
        self.total_fees_btc = 0.0
        self.cumulative_profit_usdc = 0.0
        self.cumulative_loss_usdc = 0.0
        self.last_trade_price = None
        self.backtesting_mode = False
        self.backtest_prices = []
        self.backtest_index = 0
        self.backtest_sleep = 0.0

    def append_order_history(self, text):
        self.order_history.insert(0, f"[{datetime.now().strftime('%H:%M:%S')}] {text}")
        self.order_history = self.order_history[:100]
        print(text)

    def ascii_grid(self):
        RED, GREEN, RESET = "\033[91m", "\033[92m", "\033[0m"
        lines_terminal = []
        lines_log = []
        for p in sorted(GRID_LINES, reverse=True):
            typ = "BUY " if p <= TRIGGER_PRICE else "SELL"
            symb = "○"
            color = GREEN if "BUY" in typ else RED
            lines_terminal.append(f"{p:>8.2f} ───── {color}{typ}{RESET} {symb}")
            lines_log.append(f"{p:>8.2f} ───── {typ} {symb}")
        print("📈 ASCII-GRID:")
        [print(l) for l in lines_terminal]
        self.append_order_history("📈 ASCII-GRID:")
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.order_history.extend([f"[{timestamp}] {l}" for l in lines_log])
        self.order_history = self.order_history[:100]

    def place_order(self, side, price=None, qty=None, order_type=ORDER_TYPE_LIMIT):
        try:
            safe_qty = fmt_qty_safe(qty)
            if safe_qty == 0:
                self.append_order_history("❌ Ordermenge zu klein nach Rundung, Order nicht gesetzt")
                return
            if IS_LIVE and not self.backtesting_mode:
                order = None
                if order_type == ORDER_TYPE_LIMIT:
                    order = api_call(self.client.create_order, symbol=PAIR, side=side, type=order_type,
                                     timeInForce=TIME_IN_FORCE_GTC, quantity=safe_qty, price=f"{price:.2f}")
                elif order_type == ORDER_TYPE_MARKET:
                    order = api_call(
                        (self.client.order_market_buy if side == SIDE_BUY else self.client.order_market_sell),
                        symbol=PAIR, quantity=safe_qty)
                if order and isinstance(order, dict):
                    self.open_orders[order['orderId']] = order
                self.append_order_history(
                    f"📡 LIVE-Order: {side} {order_type} @ {price if price else 'MARKET'} qty={safe_qty}")
            else:
                self.append_order_history(
                    f"🧪 SIM-Order: {side} {order_type} @ {price if price else 'MARKET'} qty={safe_qty}")
                self._simulate_order(side, price if price else (self.last_price if self.last_price else MID_PRICE),
                                     safe_qty)
        except Exception as e:
            self.append_order_history(f"Fehler beim Order setzen: {e}")

    def _simulate_order(self, side, price, qty):
        fee = qty * self.fee_rate
        if side == SIDE_BUY:
            cost = qty * price * (1 + self.fee_rate)
            if self.usdc_balance >= cost:
                self.usdc_balance -= cost
                self.btc_balance += qty - fee
                self.positions_btc += qty
                self.total_invested_usdc += cost
                self.total_fees_btc += fee
                self.append_order_history(f"SIM Buy: {qty} BTC @ {price:.2f} USDC, Gebühren {fee:.6f} BTC")
                self.last_trade_price = price
            else:
                self.append_order_history("❌ Nicht genug USDC für simulierten Kauf")
        elif side == SIDE_SELL:
            if self.btc_balance >= qty:
                proceeds = qty * price * (1 - self.fee_rate)
                cost_basis = qty * self.last_trade_price * (1 + self.fee_rate) if self.last_trade_price else proceeds
                profit = proceeds - cost_basis
                self.usdc_balance += proceeds
                self.btc_balance -= qty + fee
                self.total_fees_btc += fee
                self.net_profit_usdc += profit
                self.positions_btc -= qty
                self.total_invested_usdc -= cost_basis
                self.append_order_history(
                    f"SIM Sell: {qty} BTC @ {price:.2f} USDC, Gewinn {profit:.6f} USDC, Gebühren {fee:.6f} BTC")
                if profit >= 0:
                    self.cumulative_profit_usdc += profit
                else:
                    self.cumulative_loss_usdc += abs(profit)
                self.last_trade_price = None
            else:
                self.append_order_history("❌ Nicht genug BTC für simulierten Verkauf")

    def cancel_all_orders(self):
        try:
            if IS_LIVE and not self.backtesting_mode:
                open_orders = api_call(self.client.get_open_orders, symbol=PAIR)
                for o in open_orders:
                    api_call(self.client.cancel_order, symbol=PAIR, orderId=o['orderId'])
                self.append_order_history("Alle offenen Orders storniert (LIVE)")
                self.open_orders.clear()
            else:
                self.append_order_history("🧪 SIM: Alle Orders storniert")
        except Exception as e:
            self.append_order_history(f"Fehler Cancel Orders: {e}")

    def export_grid_overview(self):
        headers = ["Level", "Preis_USD", "Ordertyp", "Menge_BTC", "Beschreibung"]
        rows = []
        for i, p in enumerate(GRID_LINES, start=1):
            otype, desc = ("BUY", "Kauf-Linie") if p <= TRIGGER_PRICE else ("SELL", "Verkaufs-Linie")
            rows.append([i, f"{p:.2f}", otype, POSITION_SIZE_BTC, desc])
        with open("grid_overview.csv", "w", newline="", encoding='utf-8') as f:
            csv.writer(f).writerow(headers)
            csv.writer(f).writerows(rows)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grid"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        for cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            color = "C6EFCE" if cells[2].value == "BUY" else "FFC7CE"
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for c in cells:
                c.fill = fill
        wb.save("grid_overview.xlsx")
        self.append_order_history("✅ CSV- und XLSX-Datei gespeichert")

    def print_balances_and_requirements(self):
        base_asset, quote_asset = PAIR[:-4], PAIR[-4:]
        base_balance_free = base_balance_locked = 0.0
        quote_balance_free = quote_balance_locked = 0.0
        try:
            base_info = api_call(self.client.get_asset_balance, asset=base_asset) if (
                    IS_LIVE and not self.backtesting_mode) else None
            quote_info = api_call(self.client.get_asset_balance, asset=quote_asset) if (
                    IS_LIVE and not self.backtesting_mode) else None
            if base_info:
                base_balance_free = float(base_info['free'])
                base_balance_locked = float(base_info['locked'])
            if quote_info:
                quote_balance_free = float(quote_info['free'])
                quote_balance_locked = float(quote_info['locked'])
        except Exception as e:
            self.append_order_history(f"⚠ Fehler beim Abrufen der Bilanzen: {e}")

        needed_quote = POSITION_SIZE_BTC * (self.last_price if self.last_price else MID_PRICE)

        self.append_order_history(f"📊 Benötigtes Quote-Guthaben ({quote_asset}) für Trade: {needed_quote:.6f}")
        self.append_order_history(
            f"📊 Verfügbares Base-Guthaben ({base_asset}) free: {base_balance_free:.6f}, locked: {base_balance_locked:.6f}")
        self.append_order_history(
            f"📊 Verfügbares Quote-Guthaben ({quote_asset}) free: {quote_balance_free:.6f}, locked: {quote_balance_locked:.6f}")

    def market_buy_trigger(self):
        base_asset, quote_asset = PAIR[:-4], PAIR[-4:]
        needed_quote = POSITION_SIZE_BTC * (self.last_price if self.last_price else MID_PRICE)
        needed_quote = needed_quote * 1.02

        available_quote = None
        if IS_LIVE and not self.backtesting_mode:
            try:
                available_quote = float(api_call(self.client.get_asset_balance, asset=quote_asset)['free'])
            except Exception as e:
                self.append_order_history(f"❌ Fehler beim Abrufen des {quote_asset} Guthabens: {e}")
                return

            if available_quote < needed_quote:
                self.append_order_history(
                    f"❌ Nicht genug {quote_asset} für Kauf. Verfügbar: {available_quote:.6f}, Benötigt inkl. Gebühren: {needed_quote:.6f}")
                return

        qty = fmt_qty_safe(POSITION_SIZE_BTC)
        if qty == 0:
            self.append_order_history("❌ Kaufmenge nach Sicherheit und Step Size ist 0 – Order nicht gesetzt")
            return

        self.place_order(SIDE_BUY, qty=qty, order_type=ORDER_TYPE_MARKET)
        self.positions_btc += qty

    def process_price(self, price):
        self.last_stream_time = time.time()

        if self.last_price is None:
            self.last_price = price
            return

        low = min(self.last_price, price)
        high = max(self.last_price, price)

        if not self.bot_active and (low <= TRIGGER_PRICE <= high):
            self.bot_active = True
            self.append_order_history(f"▶ Trigger erreicht: {TRIGGER_PRICE}")
            self.market_buy_trigger()
            self.setup_real_grid_orders()
            Thread(target=self.order_monitor_loop, daemon=True).start()

        if self.bot_active:
            if price > RANGE_HIGH or price < STOP_LOSS_LINE:
                self.append_order_history("Exit-Bedingung – storniere alle offenen Orders und verkaufe alles")
                self.cancel_all_orders()
                self.sell_all()
                self.bot_active = False

        self.last_price = price

    def debug_trigger_now(self):
        self.append_order_history("⚡ DEBUG-TRIGGER ausgelöst!")
        self.bot_active = True
        self.market_buy_trigger()
        self.setup_real_grid_orders()
        Thread(target=self.order_monitor_loop, daemon=True).start()

    def restart_stream(self):
        try:
            if self.twm:
                self.twm.stop()
                self.append_order_history("🔄 WebSocket gestoppt – Reconnect in 5s…")
            time.sleep(self.reconnect_delay)
            self.twm = ThreadedWebsocketManager(api_key=API_KEY, api_secret=API_SECRET)

            # Asyncio Loop Handling
            try:
                if not hasattr(self.twm,
                               '_ThreadedApiManager__loop') or not self.twm._ThreadedApiManager__loop.is_running():
                    try:
                        self.twm._ThreadedApiManager__loop = asyncio.get_event_loop()
                    except RuntimeError:
                        nest_asyncio.apply()
                        self.twm._ThreadedApiManager__loop = asyncio.get_event_loop()
            except Exception:
                # Fallback falls das Loop-Handling fehlschlägt
                pass

            self.twm.start()
            self.twm.start_symbol_ticker_socket(self.on_message, symbol=PAIR)
            self.append_order_history("✅ WebSocket erfolgreich neu gestartet")
        except Exception as e:
            self.append_order_history(f"❌ Fehler beim Reconnect: {e}")

    def stream_watchdog(self):
        while self.running:
            if time.time() - self.last_stream_time > 15:
                self.append_order_history("⏱ Keine Daten seit 15s – WebSocket wird neu gestartet…")
                self.restart_stream()
            time.sleep(5)

    def setup_real_grid_orders(self):
        self.append_order_history("Setze alle Grid-Limitorders...")
        self.cancel_all_orders()
        qty = fmt_qty_safe(POSITION_SIZE_BTC)
        for p in GRID_LINES:
            if p < TRIGGER_PRICE:
                side = SIDE_BUY
            elif p > TRIGGER_PRICE:
                side = SIDE_SELL
            else:
                continue
            self.place_order(side, price=p, qty=qty, order_type=ORDER_TYPE_LIMIT)

    def order_monitor_loop(self):
        while self.running and self.bot_active:
            try:
                if IS_LIVE and not self.backtesting_mode:
                    open_orders = api_call(self.client.get_open_orders, symbol=PAIR)
                else:
                    open_orders = []

                open_order_ids = {o['orderId'] for o in open_orders if 'orderId' in o} if open_orders else set()
                known_order_ids = set(self.open_orders.keys())
                completed_orders = known_order_ids - open_order_ids

                for order_id in completed_orders:
                    order = self.open_orders.pop(order_id, None)
                    if not order:
                        continue
                    side = order['side']
                    price = float(order['price'])
                    qty = float(order['origQty'])
                    fee = qty * self.fee_rate

                    self.append_order_history(f"✅ Order ausgeführt: {side} {qty} @ {price}")

                    self.total_fees_btc += fee

                    if side == SIDE_BUY:
                        self.last_trade_price = price
                        self.positions_btc += qty
                        self.total_invested_usdc += price * qty * (1 + self.fee_rate)
                    elif side == SIDE_SELL:
                        if self.last_trade_price is not None:
                            trade_value = price * qty * (1 - self.fee_rate)
                            cost_basis = self.last_trade_price * qty * (1 + self.fee_rate)
                            profit = trade_value - cost_basis
                            if profit >= 0:
                                self.cumulative_profit_usdc += profit
                            else:
                                self.cumulative_loss_usdc += abs(profit)

                            self.net_profit_usdc += profit
                            self.positions_btc -= qty
                            self.total_invested_usdc -= cost_basis
                            self.last_trade_price = None

                    if side == SIDE_BUY:
                        next_price = self.get_next_grid_price(price, direction=1)
                        if next_price:
                            self.place_order(SIDE_SELL, price=next_price, qty=qty, order_type=ORDER_TYPE_LIMIT)
                            self.append_order_history(f"➡ Nach Kauf eine Sell-Limitorder bei {next_price} gesetzt")
                    elif side == SIDE_SELL:
                        next_price = self.get_next_grid_price(price, direction=-1)
                        if next_price:
                            self.place_order(SIDE_BUY, price=next_price, qty=qty, order_type=ORDER_TYPE_LIMIT)
                            self.append_order_history(f"➡ Nach Verkauf eine Buy-Limitorder bei {next_price} gesetzt")

                time.sleep(ORDER_CHECK_INTERVAL)

            except Exception as e:
                error_msg = str(e)
                if "Timestamp for this request was" in error_msg or "APIError(code=-1021)" in error_msg:
                    self.append_order_history("⚠ Zeit-Fehler erkannt, synchronisiere Systemzeit...")
                    sync_windows_time()
                    time.sleep(5)
                else:
                    self.append_order_history(f"Fehler beim Order-Monitoring: {e}")

    def get_next_grid_price(self, current_price, direction=1):
        sorted_grids = sorted(GRID_LINES)
        if direction == 1:
            for p in sorted_grids:
                if p > current_price:
                    return p
        elif direction == -1:
            for p in reversed(sorted_grids):
                if p < current_price:
                    return p
        return None

    def sell_all(self):
        if self.positions_btc > 0:
            qty = fmt_qty_safe(self.positions_btc)
            self.place_order(SIDE_SELL, qty=qty, order_type=ORDER_TYPE_MARKET)
            self.positions_btc = 0.0
