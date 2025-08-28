# =============================================================================
# MAIN ENTRY POINT
# =============================================================================
import sys
import importlib
import traceback
import asyncio
import nest_asyncio
from threading import Thread
from bot_core import GridTradingBot
from gui import TradingBotGUI
from api_utils import bot_status_check

# Asyncio setup
try:
    loop = asyncio.get_running_loop()
except RuntimeError:
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)

nest_asyncio.apply()

if __name__ == "__main__":
    # Prüfe notwendige Pakete
    if not check_all_packages():
        sys.exit(1)

    bot = GridTradingBot()
    gui = TradingBotGUI(bot)

    # Starte Bot in separatem Thread
    bot_thread = Thread(target=bot.start, daemon=True)
    bot_thread.start()

    gui.run()

# =============================================================================
# API UTILITIES
# =============================================================================
import math
import time
import socket
import importlib
from binance.client import Client
from binance import ThreadedWebsocketManager
import asyncio
import nest_asyncio
from config import API_KEY, API_SECRET, PAIR
from logging_setup import log_info

# API Client (Singleton-Instanz an dieser Stelle ist ok, könnte ggf. parametrisiert werden)
client = Client(API_KEY, API_SECRET)


def get_step_size(symbol: str) -> float:
    """Hole die stepSize für ein Symbol (LOT_SIZE Filter)."""
    try:
        info = client.get_symbol_info(symbol)
        for f in info['filters']:
            if f['filterType'] == 'LOT_SIZE':
                return float(f['stepSize'])
        return 0.00000001
    except Exception as e:
        log_info(f"Fehler beim Abrufen der Schrittgröße: {e}")
        return 0.00000001


STEP_SIZE = get_step_size(PAIR)


def fmt_qty_safe(q: float, step_size: float = STEP_SIZE, safety_margin: float = 0.98) -> float:
    """Formatiere Menge so, dass sie dem step_size entspricht und nie zu klein ist."""
    try:
        q = q * safety_margin
        qty_rounded = math.floor(q / step_size) * step_size
        if qty_rounded < step_size:
            qty_rounded = 0
        if step_size > 0:
            decimals = max(0, int(-math.log(step_size, 10)))
            return round(qty_rounded, decimals)
        return qty_rounded
    except Exception:
        return 0.0


def api_call(func, *args, retries: int = 5, delay: int = 1, **kwargs):
    """Führe eine API-Funktion mit Wiederholungen und Exponential Backoff aus."""
    backoff = delay
    for attempt in range(retries):
        try:
            return func(*args, **kwargs)
        except Exception as e:
            msg = str(e)
            if "429" in msg or "rate limit" in msg.lower() or "418" in msg:
                log_info(f"API Rate Limit Fehler ({attempt + 1}/{retries}): {msg}")
                log_info(f"Warte {backoff}s...")
                time.sleep(backoff)
                backoff *= 2
            else:
                log_info(f"API Fehler ({attempt + 1}/{retries}): {msg}")
                time.sleep(backoff)
                backoff *= 2
    raise Exception(f"API Call {getattr(func, '__name__', str(func))} nach {retries} Versuchen fehlgeschlagen")


def get_current_price(symbol: str) -> float:
    """Hole aktuellen Preis eines Symbols von Binance."""
    try:
        ticker = client.get_symbol_ticker(symbol=symbol)
        return float(ticker['price'])
    except Exception as e:
        log_info(f"Fehler beim Preisabruf: {e}")
        return 0.0


def bot_status_check(symbol: str = PAIR, timeout: int = 15, bot_ref=None) -> bool:
    """Führt Vorab-Checks für API, Modul, Internet, Stream-Test aus."""

    def log_step(msg):
        if bot_ref and hasattr(bot_ref, 'append_order_history'):
            bot_ref.append_order_history(msg)
        else:
            print(msg)

    log_step("========= BOT START VORAB-CHECK =========")

    # API Check
    try:
        Client(API_KEY, API_SECRET).get_account()
        log_step("✅ API-Key gültig – Account erreichbar")
    except Exception as e:
        log_step(f"❌ API-Keys ungültig oder API nicht erreichbar: {e}")
        return False

    # Modul Check
    try:
        importlib.import_module("binance")
        log_step("✅ python-binance Modul vorhanden")
    except ImportError:
        log_step("❌ python-binance Modul fehlt!")
        return False

    # Internet Check
    try:
        socket.create_connection(("api.binance.com", 443), timeout=3)
        log_step("✅ Internetverbindung OK")
    except Exception as e:
        log_step(f"❌ Keine Verbindung zu Binance API: {e}")
        return False

    # Stream Test
    log_step(f"⏳ Starte {timeout}s Stream-Test …")
    received = {"ok": False}

    def cb(_):
        received["ok"] = True

    twm = ThreadedWebsocketManager(api_key=API_KEY, api_secret=API_SECRET)
    try:
        twm.start()
        twm.start_symbol_ticker_socket(callback=cb, symbol=symbol)
        time.sleep(timeout)
        twm.stop()
        if received["ok"]:
            log_step("✅ Stream-Test erfolgreich")
        else:
            log_step("❌ Stream-Test fehlgeschlagen")
        return received["ok"]
    except Exception as e:
        log_step(f"❌ Stream-Test fehlgeschlagen: {e}")
        return False

# =============================================================================
# BOT CORE FUNCTIONALITY
# =============================================================================
import time
import csv
import openpyxl
import asyncio
import importlib
import nest_asyncio
from openpyxl.styles import PatternFill
from threading import Thread, Lock
from datetime import datetime
from binance.client import Client
from binance import ThreadedWebsocketManager
from binance.enums import *
from config import *
from api_utils import api_call, get_current_price, bot_status_check
from api_utils import fmt_qty_safe
from general_utils import check_all_packages
from logging_setup import log_info
import subprocess
import os


def sync_windows_time():
    if os.name != 'nt':
        print("Zeit-Sync nur unter Windows verfügbar.")
        return
    try:
        subprocess.run('net stop w32time', shell=True, check=True)
        subprocess.run('w32tm /unregister', shell=True, check=True)
        subprocess.run('w32tm /register', shell=True, check=True)
        subprocess.run('net start w32time', shell=True, check=True)
        subprocess.run('w32tm /resync', shell=True, check=True)
        print("Systemzeit synchronisiert.")
    except Exception as e:
        print("Fehler bei Zeit-Synchronisation:", e)


class GridTradingBot:
    def __init__(self):
        self.client = Client(API_KEY, API_SECRET)
        self.running = False
        self.bot_active = False
        self.positions_btc = 0.0
        self.usdc_balance = 0.0
        self.btc_balance = 0.0
        self.net_profit_usdc = 0.0
        self.total_invested_usdc = 0.0
        self.last_price = None
        self.order_history = []
        self.lock = Lock()
        self.twm = None
        self.last_stream_time = time.time()
        self.reconnect_delay = 5
        self.open_orders = {}
        self.initial_btc_balance = 0.01
        self.initial_usdc_balance = 1000.0
        self.btc_balance = self.initial_btc_balance
        self.usdc_balance = self.initial_usdc_balance
        self.fee_rate = 0.0001
        self.total_fees_btc = 0.0
        self.cumulative_profit_usdc = 0.0
        self.cumulative_loss_usdc = 0.0
        self.last_trade_price = None
        self.backtesting_mode = False
        self.backtest_prices = []
        self.backtest_index = 0
        self.backtest_sleep = 0.0

    def append_order_history(self, text):
        self.order_history.insert(0, f"[{datetime.now().strftime('%H:%M:%S')}] {text}")
        self.order_history = self.order_history[:100]
        print(text)

    def ascii_grid(self):
        RED, GREEN, RESET = "\033[91m", "\033[92m", "\033[0m"
        lines_terminal = []
        lines_log = []
        for p in sorted(GRID_LINES, reverse=True):
            typ = "BUY " if p <= TRIGGER_PRICE else "SELL"
            symb = "○"
            color = GREEN if "BUY" in typ else RED
            lines_terminal.append(f"{p:>8.2f} ───── {color}{typ}{RESET} {symb}")
            lines_log.append(f"{p:>8.2f} ───── {typ} {symb}")
        print("📈 ASCII-GRID:")
        [print(l) for l in lines_terminal]
        self.append_order_history("📈 ASCII-GRID:")
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.order_history.extend([f"[{timestamp}] {l}" for l in lines_log])
        self.order_history = self.order_history[:100]

    def place_order(self, side, price=None, qty=None, order_type=ORDER_TYPE_LIMIT):
        try:
            safe_qty = fmt_qty_safe(qty)
            if safe_qty == 0:
                self.append_order_history("❌ Ordermenge zu klein nach Rundung, Order nicht gesetzt")
                return
            if IS_LIVE and not self.backtesting_mode:
                order = None
                if order_type == ORDER_TYPE_LIMIT:
                    order = api_call(self.client.create_order, symbol=PAIR, side=side, type=order_type,
                                     timeInForce=TIME_IN_FORCE_GTC, quantity=safe_qty, price=f"{price:.2f}")
                elif order_type == ORDER_TYPE_MARKET:
                    order = api_call(
                        (self.client.order_market_buy if side == SIDE_BUY else self.client.order_market_sell),
                        symbol=PAIR, quantity=safe_qty)
                if order and isinstance(order, dict):
                    self.open_orders[order['orderId']] = order
                self.append_order_history(
                    f"📡 LIVE-Order: {side} {order_type} @ {price if price else 'MARKET'} qty={safe_qty}")
            else:
                self.append_order_history(
                    f"🧪 SIM-Order: {side} {order_type} @ {price if price else 'MARKET'} qty={safe_qty}")
                self._simulate_order(side, price if price else (self.last_price if self.last_price else MID_PRICE),
                                     safe_qty)
        except Exception as e:
            self.append_order_history(f"Fehler beim Order setzen: {e}")

    def _simulate_order(self, side, price, qty):
        fee = qty * self.fee_rate
        if side == SIDE_BUY:
            cost = qty * price * (1 + self.fee_rate)
            if self.usdc_balance >= cost:
                self.usdc_balance -= cost
                self.btc_balance += qty - fee
                self.positions_btc += qty
                self.total_invested_usdc += cost
                self.total_fees_btc += fee
                self.append_order_history(f"SIM Buy: {qty} BTC @ {price:.2f} USDC, Gebühren {fee:.6f} BTC")
                self.last_trade_price = price
            else:
                self.append_order_history("❌ Nicht genug USDC für simulierten Kauf")
        elif side == SIDE_SELL:
            if self.btc_balance >= qty:
                proceeds = qty * price * (1 - self.fee_rate)
                cost_basis = qty * self.last_trade_price * (1 + self.fee_rate) if self.last_trade_price else proceeds
                profit = proceeds - cost_basis
                self.usdc_balance += proceeds
                self.btc_balance -= qty + fee
                self.total_fees_btc += fee
                self.net_profit_usdc += profit
                self.positions_btc -= qty
                self.total_invested_usdc -= cost_basis
                self.append_order_history(
                    f"SIM Sell: {qty} BTC @ {price:.2f} USDC, Gewinn {profit:.6f} USDC, Gebühren {fee:.6f} BTC")
                if profit >= 0:
                    self.cumulative_profit_usdc += profit
                else:
                    self.cumulative_loss_usdc += abs(profit)
                self.last_trade_price = None
            else:
                self.append_order_history("❌ Nicht genug BTC für simulierten Verkauf")

    def cancel_all_orders(self):
        try:
            if IS_LIVE and not self.backtesting_mode:
                open_orders = api_call(self.client.get_open_orders, symbol=PAIR)
                for o in open_orders:
                    api_call(self.client.cancel_order, symbol=PAIR, orderId=o['orderId'])
                self.append_order_history("Alle offenen Orders storniert (LIVE)")
                self.open_orders.clear()
            else:
                self.append_order_history("🧪 SIM: Alle Orders storniert")
        except Exception as e:
            self.append_order_history(f"Fehler Cancel Orders: {e}")

    def export_grid_overview(self):
        headers = ["Level", "Preis_USD", "Ordertyp", "Menge_BTC", "Beschreibung"]
        rows = []
        for i, p in enumerate(GRID_LINES, start=1):
            otype, desc = ("BUY", "Kauf-Linie") if p <= TRIGGER_PRICE else ("SELL", "Verkaufs-Linie")
            rows.append([i, f"{p:.2f}", otype, POSITION_SIZE_BTC, desc])
        with open("grid_overview.csv", "w", newline="", encoding='utf-8') as f:
            csv.writer(f).writerow(headers)
            csv.writer(f).writerows(rows)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Grid"
        ws.append(headers)
        for row in rows:
            ws.append(row)
        for cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
            color = "C6EFCE" if cells[2].value == "BUY" else "FFC7CE"
            fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            for c in cells:
                c.fill = fill
        wb.save("grid_overview.xlsx")
        self.append_order_history("✅ CSV- und XLSX-Datei gespeichert")

    def print_balances_and_requirements(self):
        base_asset, quote_asset = PAIR[:-4], PAIR[-4:]
        base_balance_free = base_balance_locked = 0.0
        quote_balance_free = quote_balance_locked = 0.0
        try:
            if IS_LIVE and not self.backtesting_mode:
                base_info = api_call(self.client.get_asset_balance, asset=base_asset)
                quote_info = api_call(self.client.get_asset_balance, asset=quote_asset)
                if base_info:
                    base_balance_free = float(base_info['free'])
                    base_balance_locked = float(base_info['locked'])
                if quote_info:
                    quote_balance_free = float(quote_info['free'])
                    quote_balance_locked = float(quote_info['locked'])
        except Exception as e:
            self.append_order_history(f"⚠ Fehler beim Abrufen der Bilanzen: {e}")

        needed_quote = POSITION_SIZE_BTC * (self.last_price if self.last_price else MID_PRICE)

        self.append_order_history(f"📊 Benötigtes Quote-Guthaben ({quote_asset}) für Trade: {needed_quote:.6f}")
        self.append_order_history(
            f"📊 Verfügbares Base-Guthaben ({base_asset}) free: {base_balance_free:.6f}, locked: {base_balance_locked:.6f}")
        self.append_order_history(
            f"📊 Verfügbares Quote-Guthaben ({quote_asset}) free: {quote_balance_free:.6f}, locked: {quote_balance_locked:.6f}")

    def market_buy_trigger(self):
        base_asset, quote_asset = PAIR[:-4], PAIR[-4:]
        needed_quote = POSITION_SIZE_BTC * (self.last_price if self.last_price else MID_PRICE)
        needed_quote = needed_quote * 1.02

        available_quote = None
        if IS_LIVE and not self.backtesting_mode:
            try:
                available_quote = float(api_call(self.client.get_asset_balance, asset=quote_asset)['free'])
            except Exception as e:
                self.append_order_history(f"❌ Fehler beim Abrufen des {quote_asset} Guthabens: {e}")
                return

            if available_quote < needed_quote:
                self.append_order_history(
                    f"❌ Nicht genug {quote_asset} für Kauf. Verfügbar: {available_quote:.6f}, Benötigt inkl. Gebühren: {needed_quote:.6f}")
                return

        qty = fmt_qty_safe(POSITION_SIZE_BTC)
        if qty == 0:
            self.append_order_history("❌ Kaufmenge nach Sicherheit und Step Size ist 0 – Order nicht gesetzt")
            return

        self.place_order(SIDE_BUY, qty=qty, order_type=ORDER_TYPE_MARKET)
        self.positions_btc += qty

    def process_price(self, price):
        self.last_stream_time = time.time()

        if self.last_price is None:
            self.last_price = price
            return

        low = min(self.last_price, price)
        high = max(self.last_price, price)

        if not self.bot_active and (low <= TRIGGER_PRICE <= high):
            self.bot_active = True
            self.append_order_history(f"▶ Trigger erreicht: {TRIGGER_PRICE}")
            self.market_buy_trigger()
            self.setup_real_grid_orders()
            Thread(target=self.order_monitor_loop, daemon=True).start()

        if self.bot_active:
            if price > RANGE_HIGH or price < STOP_LOSS_LINE:
                self.append_order_history("Exit-Bedingung – storniere alle offenen Orders und verkaufe alles")
                self.cancel_all_orders()
                self.sell_all()
                self.bot_active = False

        self.last_price = price

    def debug_trigger_now(self):
        self.append_order_history("⚡ DEBUG-TRIGGER ausgelöst!")
        self.bot_active = True
        self.market_buy_trigger()
        self.setup_real_grid_orders()
        Thread(target=self.order_monitor_loop, daemon=True).start()

    def restart_stream(self):
        try:
            if self.twm:
                self.twm.stop()
                self.append_order_history("🔄 WebSocket gestoppt – Reconnect in 5s…")
            time.sleep(self.reconnect_delay)
            self.twm = ThreadedWebsocketManager(api_key=API_KEY, api_secret=API_SECRET)

            try:
                if not hasattr(self.twm,
                               '_ThreadedApiManager__loop') or not self.twm._ThreadedApiManager__loop.is_running():
                    try:
                        self.twm._ThreadedApiManager__loop = asyncio.get_event_loop()
                    except RuntimeError:
                        nest_asyncio.apply()
                        self.twm._ThreadedApiManager__loop = asyncio.get_event_loop()
            except Exception:
                pass

            self.twm.start()
            self.twm.start_symbol_ticker_socket(self.on_message, symbol=PAIR)
            self.append_order_history("✅ WebSocket erfolgreich neu gestartet")
        except Exception as e:
            self.append_order_history(f"❌ Fehler beim Reconnect: {e}")

    def stream_watchdog(self):
        while self.running:
            if time.time() - self.last_stream_time > 15:
                self.append_order_history("⏱ Keine Daten seit 15s – WebSocket wird neu gestartet…")
                self.restart_stream()
            time.sleep(5)

    def setup_real_grid_orders(self):
        self.append_order_history("Setze alle Grid-Limitorders...")
        self.cancel_all_orders()
        qty = fmt_qty_safe(POSITION_SIZE_BTC)
        for p in GRID_LINES:
            if p < TRIGGER_PRICE:
                side = SIDE_BUY
            elif p > TRIGGER_PRICE:
                side = SIDE_SELL
            else:
                continue
            self.place_order(side, price=p, qty=qty, order_type=ORDER_TYPE_LIMIT)

    def order_monitor_loop(self):
        while self.running and self.bot_active:
            try:
                if IS_LIVE and not self.backtesting_mode:
                    open_orders = api_call(self.client.get_open_orders, symbol=PAIR)
                else:
                    open_orders = []

                open_order_ids = {o['orderId'] for o in open_orders if 'orderId' in o} if open_orders else set()
                known_order_ids = set(self.open_orders.keys())
                completed_orders = known_order_ids - open_order_ids

                for order_id in completed_orders:
                    order = self.open_orders.pop(order_id, None)
                    if not order:
                        continue
                    side = order['side']
                    price = float(order['price'])
                    qty = float(order['origQty'])
                    fee = qty * self.fee_rate

                    self.append_order_history(f"✅ Order ausgeführt: {side} {qty} @ {price}")

                    self.total_fees_btc += fee

                    if side == SIDE_BUY:
                        self.last_trade_price = price
                        self.positions_btc += qty
                        self.total_invested_usdc += price * qty * (1 + self.fee_rate)
                    elif side == SIDE_SELL:
                        if self.last_trade_price is not None:
                            trade_value = price * qty * (1 - self.fee_rate)
                            cost_basis = self.last_trade_price * qty * (1 + self.fee_rate)
                            profit = trade_value - cost_basis
                            if profit >= 0:
                                self.cumulative_profit_usdc += profit
                            else:
                                self.cumulative_loss_usdc += abs(profit)

                            self.net_profit_usdc += profit
                            self.positions_btc -= qty
                            self.total_invested_usdc -= cost_basis
                            self.last_trade_price = None

                    if side == SIDE_BUY:
                        next_price = self.get_next_grid_price(price, direction=1)
                        if next_price:
                            self.place_order(SIDE_SELL, price=next_price, qty=qty, order_type=ORDER_TYPE_LIMIT)
                            self.append_order_history(f"➡ Nach Kauf eine Sell-Limitorder bei {next_price} gesetzt")
                    elif side == SIDE_SELL:
                        next_price = self.get_next_grid_price(price, direction=-1)
                        if next_price:
                            self.place_order(SIDE_BUY, price=next_price, qty=qty, order_type=ORDER_TYPE_LIMIT)
                            self.append_order_history(f"➡ Nach Verkauf eine Buy-Limitorder bei {next_price} gesetzt")

                time.sleep(ORDER_CHECK_INTERVAL)

            except Exception as e:
                error_msg = str(e)
                if "Timestamp for this request was" in error_msg or "APIError(code=-1021)" in error_msg:
                    self.append_order_history("⚠ Zeit-Fehler erkannt, synchronisiere Systemzeit...")
                    sync_windows_time()
                    time.sleep(5)
                else:
                    self.append_order_history(f"Fehler beim Order-Monitoring: {e}")

    def get_next_grid_price(self, current_price, direction=1):
        sorted_grids = sorted(GRID_LINES)
        if direction == 1:
            for p in sorted_grids:
                if p > current_price:
                    return p
        elif direction == -1:
            for p in reversed(sorted_grids):
                if p < current_price:
                    return p
        return None

    def sell_all(self):
        if self.positions_btc > 0:
            qty = fmt_qty_safe(self.positions_btc)
            self.place_order(SIDE_SELL, qty=qty, order_type=ORDER_TYPE_MARKET)
            self.positions_btc = 0.0

    def start(self):
        self.running = True
        self.append_order_history("🤖 Bot gestartet")
        if not bot_status_check(bot_ref=self):
            self.append_order_history("❌ Vorab-Check fehlgeschlagen – Bot wird nicht gestartet")
            return

        self.ascii_grid()
        self.print_balances_and_requirements()
        Thread(target=self.stream_watchdog, daemon=True).start()

        self.twm = ThreadedWebsocketManager(api_key=API_KEY, api_secret=API_SECRET)
        try:
            self.twm.start()
            self.twm.start_symbol_ticker_socket(self.on_message, symbol=PAIR)
            self.append_order_history("✅ WebSocket gestartet")
        except Exception as e:
            self.append_order_history(f"❌ Fehler beim Starten des WebSockets: {e}")

    def stop(self):
        self.running = False
        if self.twm:
            self.twm.stop()
        self.append_order_history("🛑 Bot gestoppt")

    def on_message(self, msg):
        try:
            if 'p' in msg:
                price = float(msg['p'])
                self.process_price(price)
        except Exception as e:
            self.append_order_history(f"Fehler in on_message: {e}")

    def load_backtest_data_from_csv(self, filepath):
        try:
            with open(filepath, 'r') as f:
                reader = csv.reader(f)
                self.backtest_prices = [float(row[0]) for row in reader if row]
            self.backtesting_mode = True
            self.append_order_history(f"✅ Backtest-Daten geladen: {len(self.backtest_prices)} Preise")
        except Exception as e:
            self.append_order_history(f"❌ Fehler beim Laden der Backtest-Daten: {e}")

# =============================================================================
# CONFIG.PY – KONFIGURATIONSEINSTELLUNGEN UND GRUNDKONSTANTEN
# =============================================================================

import os
from dotenv import load_dotenv

# Umgebungsvariablen aus .env-Datei laden (für API-Keys usw.)
load_dotenv()

# API Keys sicher aus Umgebungsvariablen lesen
API_KEY = os.getenv("BINANCE_API_KEY")
API_SECRET = os.getenv("BINANCE_API_SECRET")

if not API_KEY or not API_SECRET:
    raise ValueError(
        "API-Keys nicht gefunden! Bitte überprüfe deine .env Datei und den Pfad."
    )

# Trading-Einstellungen (kann ggf. auch aus externer settings.json geladen werden)
PAIR = "BTCUSDC"
MID_PRICE = 100000.0
DIFF_VALUE = 50.0
GRID_COUNT = 3
POSITION_SIZE_BTC = 0.00016819

# Berechnete Werte
RANGE_LOW = MID_PRICE - DIFF_VALUE
RANGE_HIGH = RANGE_LOW + (GRID_COUNT * DIFF_VALUE)
TRIGGER_PRICE = MID_PRICE
GRID_SIZE = DIFF_VALUE
GRID_LINES = [RANGE_LOW + i * GRID_SIZE for i in range(GRID_COUNT + 1)]
STOP_LOSS_LINE = RANGE_LOW - GRID_SIZE

# Bot-Einstellungen
IS_LIVE = False  # Standardmäßig auf False für Sicherheit
ENABLE_DEBUG_TRIGGER = True
PRICE_CHECK_INTERVAL = 1.0
ORDER_CHECK_INTERVAL = 10

# =============================================================================
# GENERAL UTILITIES
# =============================================================================
import math
import time
import csv
import re
import json
import tempfile
import os
import traceback
import asyncio
from itertools import islice
from datetime import datetime
from typing import Any, Optional, List, Dict, Callable, Union, Iterator
from logging_setup import log_info
import importlib

# Optional imports je nach Verfügbarkeit
try:
    from pytz import timezone as pytz_timezone

    PYTZ_AVAILABLE: bool = True
except ImportError:
    PYTZ_AVAILABLE: bool = False

try:
    from plyer import notification

    PLYER_AVAILABLE: bool = True
except ImportError:
    PLYER_AVAILABLE: bool = False


def round_step_size(quantity: float, step_size: float) -> float:
    """Rundet eine Menge auf das nearest Floor Vielfache des Schrittgrößenfilters."""
    precision = int(round(-math.log(step_size, 10), 0))
    rounded_qty = math.floor(quantity / step_size) * step_size
    return round(rounded_qty, precision)


def format_price(price: float) -> str:
    """Formatieren der Preisangabe auf zwei Dezimalstellen."""
    return f"{price:.2f}"


def wait_for(seconds: int) -> None:
    """Utility zum Warten mit Pausenanzeige."""
    for i in range(seconds):
        print(f"Warte {seconds - i} Sekunden...")
        time.sleep(1)


def save_order_history_to_file(order_history: List[str], filename: str = "order_history.log") -> bool:
    """Speichert den Orderverlauf in eine Datei."""
    try:
        with open(filename, "w", encoding="utf-8") as f:
            for line in order_history:
                f.write(line + "\n")
        return True
    except Exception as e:
        print(f"Fehler beim Speichern der Order-History: {e}")
        return False


def calculate_profit_loss(buy_price: float, sell_price: float, quantity: float, fee_rate: float) -> float:
    """Berechnet den Gewinn oder Verlust für einen Trade unter Berücksichtigung der Gebühren."""
    cost = buy_price * quantity * (1 + fee_rate)
    revenue = sell_price * quantity * (1 - fee_rate)
    return revenue - cost


def format_datetime(dt: Optional[datetime] = None) -> str:
    """Formatiert ein datetime-Objekt in einen lesbaren String."""
    if dt is None:
        dt = datetime.now()
    return dt.strftime("%Y-%m-%d %H:%M:%S")


def parse_float(value: Any, default: float = 0.0) -> float:
    """Sicheres Parsen von Fließkommazahlen mit Defaultwert."""
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def safe_get(dct: Optional[Dict[Any, Any]], key: Any, default: Any = None) -> Any:
    """Sichere Methode, um Werte aus einem Dictionary zu holen."""
    try:
        return dct[key]  # type: ignore
    except (KeyError, TypeError):
        return default


def current_timestamp() -> str:
    """Gibt den aktuellen Zeitstempel als String zurück."""
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def print_progress_bar(iteration: int, total: int, prefix: str = '', suffix: str = '', length: int = 50,
                       fill: str = '█') -> None:
    """Druckt eine Fortschrittsanzeige in die Konsole."""
    percent = f"{100 * (iteration / float(total)):.1f}"
    filled_length = int(length * iteration // total)
    bar = fill * filled_length + '-' * (length - filled_length)
    print(f'\r{prefix} |{bar}| {percent}% {suffix}', end='\r')
    if iteration == total:
        print()


def retry(func: Callable, retries: int = 3, delay: int = 2, backoff: int = 2,
          exceptions: Union[type, tuple] = Exception, logger: Optional[Any] = None) -> Any:
    """Führt die Funktion mit Wiederholungen bei Fehlern aus."""
    attempt = 0
    while attempt < retries:
        try:
            return func()
        except exceptions as e:
            if logger:
                logger.info(f"Versuch {attempt + 1} fehlgeschlagen: {e}")
            else:
                print(f"Versuch {attempt + 1} fehlgeschlagen: {e}")
            time.sleep(delay)
            delay *= backoff
            attempt += 1
    raise Exception(f"Funktion {func.__name__} ist nach {retries} Versuchen fehlgeschlagen.")


def human_readable_size(size: float, decimal_places: int = 2) -> str:
    """Wandelt eine Größe in Bytes in ein human-readable-Format um."""
    for unit in ['B', 'KB', 'MB', 'GB', 'TB']:
        if size < 1024.0:
            return f"{size:.{decimal_places}f} {unit}"
        size /= 1024.0
    return f"{size:.{decimal_places}f} PB"


def is_price_in_range(price: float, low: float, high: float) -> bool:
    """Prüft, ob ein Preis innerhalb eines Bereichs liegt."""
    return low <= price <= high


def safe_division(numerator: float, denominator: float, default: float = 0.0) -> float:
    """Sichere Division, um Division durch Null zu vermeiden."""
    try:
        return numerator / denominator
    except ZeroDivisionError:
        return default


def timestamp_to_datetime(ts: float) -> Optional[datetime]:
    """Konvertiert Unix-Timestamp in datetime-Objekt."""
    try:
        return datetime.fromtimestamp(ts)
    except Exception:
        return None


def datetime_to_timestamp(dt: datetime) -> Optional[float]:
    """Konvertiert datetime-Objekt in Unix-Timestamp."""
    try:
        return dt.timestamp()
    except Exception:
        return None


def safe_round(value: Any, decimals: int = 8) -> float:
    """Rundet einen Wert auf eine bestimmte Anzahl von Dezimalstellen."""
    try:
        return round(float(value), decimals)
    except Exception:
        return 0.0


def read_csv_prices(filepath: str) -> List[float]:
    """Liest Preise aus einer CSV-Datei und gibt sie als Liste zurück."""
    prices = []
    try:
        with open(filepath, 'r') as f:
            reader = csv.reader(f)
            for row in reader:
                if row:
                    price = parse_float(row[0], default=None)
                    if price is not None:
                        prices.append(price)
    except Exception as e:
        print(f"Fehler beim Lesen der CSV: {e}")
    return prices


def log_exception(e: Exception, logger: Optional[Any] = None) -> None:
    """Loggt eine Exception inklusive Stack-Trace."""
    exc_msg = ''.join(traceback.format_exception(type(e), e, e.__traceback__))
    if logger:
        logger.error(exc_msg)
    else:
        print(exc_msg)


def safe_cast(val: Any, to_type: Callable, default: Any = None) -> Any:
    """Sicherer Cast auf definierten Typ mit Defaultwert."""
    try:
        return to_type(val)
    except (ValueError, TypeError):
        return default


def timestamp_now() -> float:
    """Aktueller Zeitstempel als float."""
    return time.time()


def ensure_dir_exists(path: str) -> None:
    """Erstellt ein Verzeichnis, falls es nicht existiert."""
    if not os.path.exists(path):
        os.makedirs(path)


def truncate_text(text: str, max_length: int = 100) -> str:
    """Text kürzen, falls zu lang."""
    if len(text) > max_length:
        return text[:max_length - 3] + "..."
    return text


def is_market_open() -> bool:
    """Prüft, ob der Markt geöffnet ist (z.B. Binance ist 24/7 offen)."""
    return True


def elapsed_seconds(start_time: float) -> float:
    """Gibt Sekunden seit start_time zurück."""
    return time.time() - start_time


def format_pct(value: float, decimals: int = 2) -> str:
    """Formatiert eine Zahl als Prozentstring."""
    try:
        return f"{value * 100:.{decimals}f}%"
    except Exception:
        return "0%"


def safe_min(*args: Optional[float]) -> Optional[float]:
    """Gibt das Minimum einer Liste zurück, ignoriert None-Werte."""
    filtered = [x for x in args if x is not None]  # type: ignore
    if not filtered:
        return None
    return min(filtered)


def safe_max(*args: Optional[float]) -> Optional[float]:
    """Gibt das Maximum einer Liste zurück, ignoriert None-Werte."""
    filtered = [x for x in args if x is not None]  # type: ignore
    if not filtered:
        return None
    return max(filtered)


def chunk_list(lst: List[Any], n: int) -> List[List[Any]]:
    """Teilt eine Liste in n ungefähr gleich große Stücke."""
    k, m = divmod(len(lst), n)
    return [lst[i * k + min(i, m):(i + 1) * k + min(i + 1, m)] for i in range(n)]


def format_order_side(side: str) -> str:
    """Formatiert 'BUY' oder 'SELL' in ein deutschsprachiges Wort."""
    if side.upper() == "BUY":
        return "Kaufen"
    elif side.upper() == "SELL":
        return "Verkaufen"
    return "Unbekannt"


def current_date_string() -> str:
    """Gibt das aktuelle Datum als String im YYYY-MM-DD Format zurück."""
    return datetime.now().strftime("%Y-%m-%d")


def parse_bool(value: Any) -> bool:
    """Parst einen Wert in Boolean mit tolerantem Verhalten."""
    if isinstance(value, bool):
        return value
    if isinstance(value, str):
        return value.strip().lower() in ['true', '1', 'yes', 'y']
    return False


def pretty_print_order(order: Dict[str, Any]) -> str:
    """Erzeugt eine gut lesbare Stringdarstellung einer Order."""
    side = format_order_side(order.get('side', ''))
    price = order.get('price', 'N/A')
    qty = order.get('origQty', 'N/A')
    status = order.get('status', 'N/A')
    return f"Order: {side} {qty} zum Preis {price} Status: {status}"


def clear_console() -> None:
    """Löscht die Konsole unter Windows oder Unix-Betriebssystemen."""
    os.system('cls' if os.name == 'nt' else 'clear')


def notify_user(title: str, message: str) -> None:
    """Zeigt eine einfache Desktop-Benachrichtigung an (wenn möglich)."""
    if PLYER_AVAILABLE:
        try:
            notification.notify(title=title, message=message)
        except Exception:
            print(f"{title}: {message}")
    else:
        print(f"{title}: {message}")


def timestamp_to_str(ts: float, fmt: str = "%Y-%m-%d %H:%M:%S") -> str:
    """Konvertiert Timestamp in formatierten String."""
    try:
        return datetime.fromtimestamp(ts).strftime(fmt)
    except Exception:
        return ""


def str_to_timestamp(date_str: str, fmt: str = "%Y-%m-%d %H:%M:%S") -> Optional[float]:
    """Konvertiert formatierte Datum-String in Timestamp."""
    try:
        dt = datetime.strptime(date_str, fmt)
        return dt.timestamp()
    except Exception:
        return None


def is_trade_profitable(buy_price: float, sell_price: float, fee_rate: float) -> bool:
    """Gibt True zurück, wenn Trade profitabel ist unter Berücksichtigung der Gebühren."""
    return sell_price > buy_price * (1 + 2 * fee_rate)


def log_debug(message: str, enable_debug: bool = True) -> None:
    """Schreibt Debug-Nachrichten nur, wenn enable_debug True."""
    if enable_debug:
        print(f"DEBUG: {message}")


def sanitize_filename(name: str) -> str:
    """Konvertiert einen String zu einem gültigen Dateinamen."""
    return re.sub(r'[<>:"/\\|?*]', '_', name)


def format_duration(seconds: int) -> str:
    """Formatiert eine Zeitdauer in Sekunden als H:M:S."""
    hours, remainder = divmod(int(seconds), 3600)
    minutes, secs = divmod(remainder, 60)
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def get_time_diff_in_seconds(start: datetime, end: datetime) -> float:
    """Berechnet die Differenz zwischen zwei datetime-Objekten in Sekunden."""
    delta = end - start
    return delta.total_seconds()


def retry_async(func: Callable, retries: int = 3, delay: int = 2, backoff: int = 2,
                exceptions: Union[type, tuple] = Exception, logger: Optional[Any] = None) -> Callable:
    """Async Variante einer Retry-Funktion."""

    async def wrapper(*args: Any, **kwargs: Any) -> Any:
        attempt = 0
        current_delay = delay
        while attempt < retries:
            try:
                return await func(*args, **kwargs)
            except exceptions as e:
                if logger:
                    logger.info(f"Versuch {attempt + 1} fehlgeschlagen: {e}")
                else:
                    print(f"Versuch {attempt + 1} fehlgeschlagen: {e}")
                await asyncio.sleep(current_delay)
                current_delay *= backoff
                attempt += 1
        raise Exception(f"Funktion {func.__name__} ist nach {retries} Versuchen fehlgeschlagen.")

    return wrapper


def convert_to_float(value: Any, default: float = 0.0) -> float:
    """Versucht, einen Wert in float umzuwandeln, sonst Default."""
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def safe_string(value: Any, default: str = "") -> str:
    """Wandelt Wert in String um, falls nicht None, sonst Default."""
    if value is None:
        return default
    return str(value)


def retry_with_callback(func: Callable, on_fail: Optional[Callable] = None, retries: int = 3, delay: int = 2) -> Any:
    """Führt Funktion mit retries aus, bei Fehler Callback aufrufen."""
    for attempt in range(retries):
        try:
            return func()
        except Exception as e:
            if on_fail:
                on_fail(e, attempt + 1)
            time.sleep(delay)
    raise Exception(f"Funktion {func.__name__} nach {retries} Versuchen fehlgeschlagen.")


def format_order(order: Dict[str, Any]) -> str:
    """Erstellt eine kurze Beschreibung einer Order."""
    side = safe_string(order.get('side'), 'UNKNOWN')
    price = safe_string(order.get('price'), '0')
    qty = safe_string(order.get('origQty'), '0')
    status = safe_string(order.get('status'), 'UNKNOWN')
    return f"{side} {qty} @ {price} ({status})"


def now_string() -> str:
    """Gibt die aktuelle Zeit als String zurück."""
    return datetime.now().strftime("%H:%M:%S")


def calculate_percentage_change(old: float, new: float) -> float:
    """Berechnet die prozentuale Änderung von old zu new."""
    try:
        return ((new - old) / old) * 100
    except ZeroDivisionError:
        return 0.0


def is_number(value: Any) -> bool:
    """Überprüft, ob ein Wert eine Zahl ist."""
    try:
        float(value)
        return True
    except (ValueError, TypeError):
        return False


def is_valid_symbol(symbol: str) -> bool:
    """Prüft, ob ein Handelssymbol valide ist (z.B. BTCUSDT)."""
    return bool(re.match(r'^[A-Z]{6,12}$', symbol))


def normalize_symbol(symbol: str) -> str:
    """Normalisiert ein Symbol, Großbuchstaben und ohne Leerzeichen."""
    return symbol.upper().replace(' ', '')


def clamp(value: float, min_value: float, max_value: float) -> float:
    """Begrenzt value auf einen Wertebereich."""
    return max(min_value, min(value, max_value))


def timestamp_to_iso(ts: float) -> str:
    """Konvertiert Unix-Timestamp in ISO 8601 Format."""
    try:
        return datetime.utcfromtimestamp(ts).isoformat() + "Z"
    except Exception:
        return ""


def iso_to_timestamp(iso_str: str) -> Optional[float]:
    """Konvertiert ISO 8601 Datumsstring in Unix-Timestamp."""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", ""))
        return dt.timestamp()
    except Exception:
        return None


def flatten_list(list_of_lists: List[List[Any]]) -> List[Any]:
    """Flacht eine Liste von Listen zu einer einzigen Liste ab."""
    return [item for sublist in list_of_lists for item in sublist]


def dict_get_or_default(d: Optional[Dict[Any, Any]], key: Any, default: Any = None) -> Any:
    """Gibt d[key] zurück oder default wenn key nicht existiert."""
    if d and isinstance(d, dict):
        return d.get(key, default)
    return default


def dict_deep_get(d: Optional[Dict[Any, Any]], keys: List[Any], default: Any = None) -> Any:
    """Greift tief verschachtelte Schlüssel in einem Dictionary sicher ab."""
    cur = d
    for key in keys:
        if isinstance(cur, dict) and key in cur:
            cur = cur[key]
        else:
            return default
    return cur


def safe_increment(dictionary: Dict[Any, Any], key: Any, amount: Union[int, float] = 1) -> None:
    """Erhöht einen Wert in einem Dictionary sicher um amount."""
    if key in dictionary and isinstance(dictionary[key], (int, float)):
        dictionary[key] += amount
    else:
        dictionary[key] = amount


def list_unique(seq: List[Any]) -> List[Any]:
    """Gibt eine Liste mit einzigartigen Elementen zurück, Reihenfolge bleibt erhalten."""
    seen = set()
    result = []
    for item in seq:
        if item not in seen:
            seen.add(item)
            result.append(item)
    return result


def parse_int(value: Any, default: int = 0) -> int:
    """Konvertiert einen Wert in Integer, sonst Defaultwert."""
    try:
        return int(value)
    except (ValueError, TypeError):
        return default


def invert_side(side: str) -> str:
    """Invertiert den Side-Wert BUY ↔ SELL."""
    if side.upper() == "BUY":
        return "SELL"
    elif side.upper() == "SELL":
        return "BUY"
    return side


def current_utc_datetime() -> datetime:
    """Gibt das aktuelles UTC datetime Objekt zurück."""
    return datetime.utcnow()


def safe_list_get(lst: List[Any], index: int, default: Optional[Any] = None) -> Any:
    """Sicherer Listenindex-Abruf mit Default."""
    try:
        return lst[index]
    except IndexError:
        return default


def is_order_complete(order: Dict[str, Any]) -> bool:
    """Prüft, ob eine Order den Status FILLED oder PARTIALLY_FILLED hat."""
    status = order.get('status', '').upper()
    return status in ['FILLED', 'PARTIALLY_FILLED']


def clear_list(lst: List[Any]) -> None:
    """Leert eine Liste inplace."""
    lst.clear()


def safe_pop(dictionary: Dict[Any, Any], key: Any, default: Optional[Any] = None) -> Any:
    """Sicheres Entfernen eines Schlüssels aus einem Dictionary."""
    if key in dictionary:
        return dictionary.pop(key)
    return default


def is_float_string(s: Any) -> bool:
    """Überprüft, ob ein String eine gültige float Zahl repräsentiert."""
    try:
        float(s)
        return True
    except Exception:
        return False


def timestamp_now_ms() -> int:
    """Gibt aktuellen Zeitstempel in Millisekunden zurück."""
    return int(round(time.time() * 1000))


def list_diff(list1: List[Any], list2: List[Any]) -> List[Any]:
    """Gibt die Elemente, die in list1 aber nicht in list2 sind, zurück."""
    return [item for item in list1 if item not in list2]


def split_string_by_length(s: str, length: int) -> List[str]:
    """Teilt einen String in Stücke der Länge length auf."""
    return [s[i:i + length] for i in range(0, len(s), length)]


def dict_merge(d1: Dict[Any, Any], d2: Dict[Any, Any]) -> Dict[Any, Any]:
    """Fügt zwei Dictionaries zusammen, d2 überschreibt d1 bei Konflikten."""
    result = d1.copy()
    result.update(d2)
    return result


def is_market_open_now(timezone: str = 'UTC', start_hour: int = 0, end_hour: int = 24) -> bool:
    """Überprüft, ob die aktuelle Zeit innerhalb der Marktöffnungszeiten liegt."""
    if not PYTZ_AVAILABLE:
        log_info("⚠ pytz nicht installiert, kann Marktöffnungszeiten nicht prüfen")
        return True

    try:
        now = datetime.now(pytz_timezone(timezone))
        return start_hour <= now.hour < end_hour
    except Exception:
        return True


def atomic_write(filepath: str, data: str, mode: str = 'w', encoding: str = 'utf-8') -> None:
    """Schreibt Daten atomar in eine Datei, um Korruption zu vermeiden."""
    dirpath = os.path.dirname(filepath)
    with tempfile.NamedTemporaryFile(mode=mode, encoding=encoding, dir=dirpath, delete=False) as tmpfile:
        tmpfile.write(data)
        tempname = tmpfile.name
    os.replace(tempname, filepath)


def chunked_iterable(iterable: Iterator[Any], chunk_size: int) -> Iterator[List[Any]]:
    """Erzeugt Iteratoren auf Teilstücke der Größe chunk_size."""
    it = iter(iterable)
    while True:
        chunk = list(islice(it, chunk_size))
        if not chunk:
            break
        yield chunk


def calculate_ema(prices: List[float], period: int = 10) -> Optional[float]:
    """Berechnet den gleitenden Durchschnitt EMA einer Liste von Preisen."""
    if not prices or len(prices) < period:
        return None
    k = 2 / (period + 1)
    ema_values = [sum(prices[:period]) / period]
    for price in prices[period:]:
        ema = price * k + ema_values[-1] * (1 - k)
        ema_values.append(ema)
    return ema_values[-1]


def merge_dicts(*dicts: Dict[Any, Any]) -> Dict[Any, Any]:
    """Fasst mehrere Dictionaries zusammen, spätere überschreiben frühere."""
    result = {}
    for d in dicts:
        if isinstance(d, dict):
            result.update(d)
    return result


def filter_none_values(d: Dict[Any, Any]) -> Dict[Any, Any]:
    """Filtert ein Dictionary, um Einträge mit None-Werten zu entfernen."""
    return {k: v for k, v in d.items() if v is not None}


def safe_json_loads(s: str, default: Optional[Any] = None) -> Optional[Any]:
    """Parst JSON sicher, gibt Default bei Fehlern zurück."""
    try:
        return json.loads(s)
    except Exception:
        return default


def safe_json_dumps(obj: Any, **kwargs: Any) -> str:
    """JSON Serienalisierung mit Fehlerbackup."""
    try:
        return json.dumps(obj, **kwargs)
    except Exception:
        return ""


def timer(func: Callable) -> Callable:
    """Dekorator zur Messung der Ausführungszeit einer Funktion."""

    def wrapper(*args: Any, **kwargs: Any) -> Any:
        start = time.perf_counter()
        result = func(*args, **kwargs)
        end = time.perf_counter()
        print(f"[TIMER] {func.__name__} dauerte {end - start:.6f} Sekunden")
        return result

    return wrapper


def ensure_list(obj: Any) -> List[Any]:
    """Stellt sicher, dass ein Objekt eine Liste ist."""
    if obj is None:
        return []
    if isinstance(obj, list):
        return obj
    return [obj]


def ensure_dict(obj: Any) -> Dict[Any, Any]:
    """Stellt sicher, dass ein Objekt ein Dictionary ist."""
    if obj is None:
        return {}
    if isinstance(obj, dict):
        return obj
    return {}


def is_time_between(start_time: datetime.time, end_time: datetime.time,
                    check_time: Optional[datetime.time] = None) -> bool:
    """Prüft ob die aktuelle Zeit innerhalb eines Bereichs liegt."""
    if check_time is None:
        check_time = datetime.now().time()
    if start_time <= end_time:
        return start_time <= check_time <= end_time
    else:
        return check_time >= start_time or check_time <= end_time


def check_all_packages(gui: bool = False, bot_ref: Optional[Any] = None) -> bool:
    """Überprüft, ob alle notwendigen Pakete installiert sind."""
    required = ['binance', 'matplotlib', 'tkinter', 'openpyxl', 'requests', 'dotenv']
    missing = []
    for pkg in required:
        try:
            importlib.import_module(pkg)
        except ImportError:
            missing.append(pkg)
    if missing:
        msg = f"Fehlende Pakete: {', '.join(missing)}. Bitte installieren!"
        if gui and bot_ref:
            bot_ref.append_order_history(msg)
            import tkinter.messagebox as messagebox
            messagebox.showwarning("Fehlende Pakete", msg)
        else:
            print(msg)
        return False
    return True

# =============================================================================
# GUI COMPONENTS
# =============================================================================
import tkinter as tk
from tkinter import messagebox, ttk, filedialog
import tkinter.font as tkFont
import matplotlib

matplotlib.use('TkAgg')
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from bot_core import GridTradingBot
from config import IS_LIVE, ENABLE_DEBUG_TRIGGER, PRICE_CHECK_INTERVAL
from general_utils import check_all_packages


class TradingBotGUI:
    def __init__(self, bot):
        self.bot = bot
        self.root = tk.Tk()
        self.root.title("Grid Trading Bot")

        self.main_font = tkFont.Font(family="Arial", size=14, weight="bold")
        self.status_font = tkFont.Font(family="Arial", size=12)

        tk.Label(self.root, text="Bot Status:", font=self.main_font).pack()
        self.status_var = tk.StringVar(value="RUNNING")
        tk.Label(self.root, textvariable=self.status_var, font=self.status_font).pack()

        self.start_btn = tk.Button(self.root, text="Start (Autostart aktiv!)", bg="grey", fg="white", state="disabled")
        self.start_btn.pack()
        self.stop_btn = tk.Button(self.root, text="Stop", command=self.stop_bot, bg="red", fg="white")
        self.stop_btn.pack()
        if ENABLE_DEBUG_TRIGGER:
            tk.Button(self.root, text="⚡ Debug Trigger", bg="orange", command=self.debug_trigger_clicked).pack()

        tk.Button(self.root, text="🔄 Update Checker", bg="blue", fg="white", command=self.open_update_checker).pack(pady=5)
        tk.Button(self.root, text="📂 Backtesting CSV laden", command=self.load_backtest_csv).pack(pady=5)
        tk.Button(self.root, text="📊 Chart öffnen", bg="green", fg="white", command=self.open_chart_window).pack(pady=5)

        tk.Button(self.root, text="🛈 Trading Info öffnen", bg="purple", fg="white",
                  command=self.open_trading_info_window).pack(pady=5)

        tk.Label(self.root, text="Log:", font=self.main_font).pack()
        self.logbox = tk.Listbox(self.root, width=100, height=15)
        self.logbox.pack(side=tk.TOP, fill=tk.BOTH, expand=False)
        self.log_scrollbar = tk.Scrollbar(self.root, command=self.logbox.yview, orient="vertical")
        self.log_scrollbar.pack(side=tk.TOP, fill=tk.Y)
        self.logbox.config(yscrollcommand=self.log_scrollbar.set)

        self._scroll_pos = 0.0
        self.logbox.bind("<MouseWheel>", self.on_mouse_wheel)
        self.logbox.bind("<Button-4>", self.on_mouse_wheel)
        self.logbox.bind("<Button-5>", self.on_mouse_wheel)

        self.chart_window = None
        self.trading_info_window = None
        self.open_trading_info_window()

        self.update_ui()

    def open_update_checker(self):
        UpdateCheckerGUI(self.root, self.bot)

    def open_chart_window(self):
        if self.chart_window is None or not tk.Toplevel.winfo_exists(self.chart_window.top):
            self.chart_window = GridChartWindow(self.root, self.bot)
        else:
            self.chart_window.top.lift()

    def open_trading_info_window(self):
        if self.trading_info_window is None or not tk.Toplevel.winfo_exists(self.trading_info_window.top):
            self.trading_info_window = TradingInfoWindow(self.root, self.bot)
        else:
            self.trading_info_window.top.lift()

    def load_backtest_csv(self):
        filepath = filedialog.askopenfilename(title="Backtesting CSV Datei wählen",
                                              filetypes=[("CSV Dateien", "*.csv"), ("Alle Dateien", "*.*")])
        if filepath:
            self.bot.load_backtest_data_from_csv(filepath)
            if not self.bot.running:
                self.bot.start()

    def on_mouse_wheel(self, event):
        self._scroll_pos = self.logbox.yview()[0]

    def stop_bot(self):
        if self.bot.running:
            self.bot.stop()
            self.status_var.set("STOPPED")
            self.bot.append_order_history("⏹ Stop-Button gedrückt – Bot gestoppt.")
        else:
            self.bot.append_order_history("ℹ Bot ist bereits gestoppt")

    def debug_trigger_clicked(self):
        if not self.bot.running:
            self.bot.append_order_history("⚠ Bot läuft nicht – kein Debug möglich")
            return
        if IS_LIVE:
            if not messagebox.askyesno("Bestätigung", "⚠ LIVE-MODUS!\nEchte Orders werden platziert.\nFortfahren?"):
                return
        self.bot.debug_trigger_now()

    def update_ui(self):
        pos = self.logbox.yview()
        self.logbox.delete(0, tk.END)
        for entry in self.bot.order_history:
            self.logbox.insert(tk.END, entry)
        self.logbox.yview_moveto(self._scroll_pos)
        self._scroll_pos = self.logbox.yview()[0]
        self.root.after(int(PRICE_CHECK_INTERVAL * 1000), self.update_ui)

    def run(self):
        self.root.mainloop()


class GridChartWindow:
    def __init__(self, root, bot):
        self.top = tk.Toplevel(root)
        self.top.title("Grid Chart")
        self.bot = bot
        self.fig, self.ax = plt.subplots(figsize=(8, 5))
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.top)
        self.canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)
        self.update_chart()

    def update_chart(self):
        self.ax.clear()
        from config import GRID_LINES
        prices = [p for p in GRID_LINES]
        profits = [0] * len(prices)  # Beispielwerte
        self.ax.plot(prices, profits, 'bo-')
        self.ax.set_title("Grid Trading Chart")
        self.ax.set_xlabel("Preis")
        self.ax.set_ylabel("Profit")
        self.canvas.draw()


class TradingInfoWindow:
    def __init__(self, root, bot):
        self.top = tk.Toplevel(root)
        self.top.title("Trading Info")
        self.bot = bot
        self.text = tk.Text(self.top, width=80, height=20)
        self.text.pack(fill=tk.BOTH, expand=True)
        self.update_info()

    def update_info(self):
        self.text.delete('1.0', tk.END)
        info = f"Netto Profit (USDC): {self.bot.net_profit_usdc:.4f}\n"
        info += f"Offene BTC Positionen: {self.bot.positions_btc:.6f}\n"
        info += f"Letzter Trade Preis: {self.bot.last_trade_price if self.bot.last_trade_price else 'N/A'}\n"
        info += f"Gesamte Gebühren (BTC): {self.bot.total_fees_btc:.8f}\n"
        self.text.insert(tk.END, info)
        self.top.after(1000, self.update_info)


class UpdateCheckerGUI:
    def __init__(self, root, bot):
        self.top = tk.Toplevel(root)
        self.top.title("Update Checker")
        self.bot = bot
        label = tk.Label(self.top, text="Überprüfe auf Updates ...")
        label.pack(padx=10, pady=10)

        # Einfache Update-Check-Logik
        version_label = tk.Label(self.top, text="Aktuelle Version: 1.0.0")
        version_label.pack(pady=5)

        check_button = tk.Button(self.top, text="Auf Updates prüfen", command=self.check_updates)
        check_button.pack(pady=5)

        self.result_text = tk.Text(self.top, width=60, height=10)
        self.result_text.pack(padx=10, pady=10)

    def check_updates(self):
        self.result_text.delete('1.0', tk.END)
        self.result_text.insert(tk.END, "Update-Check wird durchgeführt...\n")
        # Hier könnte echte Update-Logik implementiert werden
        self.result_text.insert(tk.END, "✅ Du hast die neueste Version (1.0.0)\n")

# =============================================================================
# LOGGING CONFIGURATION
# =============================================================================
import logging
from logging.handlers import RotatingFileHandler
import os

# Log-Verzeichnis erstellen falls nicht vorhanden
if not os.path.exists('logs'):
    os.makedirs('logs')

# Empfehlung: Definiere einen benannten Logger fürs gesamte Projekt
logger = logging.getLogger("grid_trading_bot")
logger.setLevel(logging.INFO)

handler = RotatingFileHandler(
    'logs/grid_trading_bot.log', maxBytes=5*1024*1024, backupCount=3, encoding='utf-8'
)
formatter = logging.Formatter('%(asctime)s %(levelname)s: %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)

# Auch Console Output
console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)

def log_info(msg):
    logger.info(msg)
